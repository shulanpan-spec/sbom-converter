#!/usr/bin/env python3
"""
SBOM Excel to SPDX 2.3 JSON Converter
Converts automotive SBOM Excel files to SPDX 2.3 compliant JSON format.
Ensures VBF package contains all other packages via CONTAINS relationships.
"""

import os
import sys
import json
import uuid
import glob
import argparse
import re
from datetime import datetime, timezone

# --- Dependency check (import only, no auto-install) ---
try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: Missing required packages. Please run:")
    print("  pip install -r requirements.txt")
    sys.exit(1)


def spdx_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def clean_vbf_part_number(vbf_number: str) -> str:
    """Normalize a VBF part number string."""
    if not vbf_number:
        return ""
    vbf_number = str(vbf_number).strip().replace(" ", "")
    vbf_number = re.sub(r'^VBF[-_]?', '', vbf_number, flags=re.IGNORECASE)
    vbf_number = re.sub(r'[^\w\d.-]', '', vbf_number)
    return vbf_number


def parse_relationship_targets(target_str: str) -> list:
    """Parse relationship targets, handling newline-separated multi-values."""
    if not target_str or pd.isna(target_str):
        return []
    return [t.strip() for t in str(target_str).strip().split('\n') if t.strip()]


def extract_sbom_from_excel(file_path: str):
    """
    Extract SBOM data from a single Excel file.
    Returns (spdx_doc dict, vbf_part_number str) or (None, None) on failure.
    """
    print(f"Processing: {os.path.basename(file_path)}")

    try:
        wb = load_workbook(file_path, data_only=True)

        # ---------- Cover sheet ----------
        if "Cover" not in wb.sheetnames:
            print(f"  WARNING: No 'Cover' sheet found in {file_path}")
            return None, None

        metadata = {}
        for row in wb["Cover"].iter_rows(values_only=True):
            if row and len(row) >= 2 and row[0] and row[1]:
                key = str(row[0]).strip()
                value = str(row[1]).strip()
                if key and value and not key.startswith("Please fill"):
                    metadata[key] = value

        if not metadata:
            print(f"  WARNING: 'Cover' sheet is empty or incorrectly formatted.")
            return None, None

        vbf_part_number = clean_vbf_part_number(
            metadata.get("VBF-part_x0002_number", "")
        ) or clean_vbf_part_number(os.path.splitext(os.path.basename(file_path))[0])

        print(f"  VBF part number: {vbf_part_number}")

        # ---------- SBOM sheet ----------
        if "SBOM" not in wb.sheetnames:
            print(f"  WARNING: No 'SBOM' sheet found in {file_path}")
            return None, vbf_part_number

        sbom_ws = wb["SBOM"]
        headers = []
        header_row = 2

        for row_num in range(1, 4):
            row_values = list(next(sbom_ws.iter_rows(
                min_row=row_num, max_row=row_num, values_only=True
            )))
            if sum(1 for v in row_values if v and str(v).strip()) >= 3:
                header_row = row_num
                headers = [
                    str(v).strip().lower().replace(" ", "").replace("_", "")
                    if v else f"col_{i}"
                    for i, v in enumerate(row_values)
                ]
                break

        if not headers:
            print(f"  ERROR: No valid header row found in SBOM sheet.")
            return None, vbf_part_number

        sbom_data = []
        for row in sbom_ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(row):
                continue
            row_dict = {
                headers[i]: v
                for i, v in enumerate(row)
                if i < len(headers) and pd.notna(v) and str(v).strip()
            }
            if row_dict:
                sbom_data.append(row_dict)

        if not sbom_data:
            print(f"  WARNING: No data rows found in SBOM sheet.")
            return None, vbf_part_number

        sbom_df = pd.DataFrame(sbom_data)
        sbom_df.columns = [
            str(c).strip().lower().replace(" ", "").replace("_", "").replace(".", "")
            for c in sbom_df.columns
        ]

        # ---------- Build packages ----------
        packages = []
        all_package_ids = set()

        for _, row in sbom_df.iterrows():
            spdx_id = str(row.get("spdxid", "")).strip()
            if not spdx_id:
                continue

            all_package_ids.add(spdx_id)

            supplier_raw = str(row.get("supplier", "")).strip() if pd.notna(row.get("supplier")) else ""
            supplier = (
                supplier_raw if supplier_raw.startswith("Organization:")
                else f"Organization: {supplier_raw}" if supplier_raw
                else ""
            )

            version_info = (
                str(row.get("versioninfo", "")).strip()
                or str(row.get("identifier", "")).strip()
            )

            pkg = {
                "SPDXID": spdx_id,
                "name": str(row.get("name", "")).strip(),
                "versionInfo": version_info,
                "supplier": supplier,
                "downloadLocation": str(row.get("downloadlocation", "NONE")).strip() or "NONE",
                "licenseDeclared": str(row.get("licensedeclared", "NONE")).strip() or "NONE",
                "copyrightText": str(row.get("copyrighttext", "")).strip(),
                "externalRefs": [],
            }

            ext_ref = str(row.get("externalrefs", "")).strip() if pd.notna(row.get("externalrefs")) else ""
            if ext_ref and ext_ref != "NONE":
                pkg["externalRefs"].append({
                    "referenceType": "OTHER",
                    "referenceLocator": ext_ref,
                })

            packages.append(pkg)

        # ---------- Build relationships (use set for O(1) duplicate check) ----------
        ROOT_PKG = "SPDXRef-Package-VBF"
        relationships = []
        existing_rels: set = set()  # (src, type, target)

        def add_relationship(src: str, rel_type: str, target: str):
            key = (src, rel_type, target)
            if key not in existing_rels and src != target:
                existing_rels.add(key)
                relationships.append({
                    "spdxElementId": src,
                    "relationshipType": rel_type,
                    "relatedSpdxElement": target,
                })

        # VBF CONTAINS all other packages
        if ROOT_PKG in all_package_ids:
            for pid in all_package_ids:
                if pid != ROOT_PKG:
                    add_relationship(ROOT_PKG, "CONTAINS", pid)
            print(f"  VBF CONTAINS relationships: {len(all_package_ids) - 1}")

        # DEPENDS_ON relationships between non-VBF packages
        for _, row in sbom_df.iterrows():
            current_id = str(row.get("spdxid", "")).strip()
            if not current_id or current_id == ROOT_PKG:
                continue

            rtype = str(row.get("relationshiptype", "")).strip().upper()
            if rtype != "DEPENDS_ON":
                continue

            targets = parse_relationship_targets(row.get("relatedspdxelement"))
            sources_raw = parse_relationship_targets(row.get("spdxelementid"))
            sources = [s for s in sources_raw if s in all_package_ids and s != ROOT_PKG] \
                      or ([current_id] if current_id in all_package_ids else [])

            for src in sources:
                for tgt in targets:
                    if tgt in all_package_ids and tgt != ROOT_PKG:
                        add_relationship(src, "DEPENDS_ON", tgt)

        print(f"  Extracted: {len(packages)} packages, {len(relationships)} relationships")

        # ---------- Assemble SPDX document ----------
        name = metadata.get("name", "")
        if not name or "B8" in name or "=" in name:
            name = f"VBF-{vbf_part_number} SPDX SBOM" if vbf_part_number \
                   else f"SBOM-{os.path.splitext(os.path.basename(file_path))[0]}"

        org_name = metadata.get(
            "Organization",
            "XXXXXXXXXXXXXXX Co., Ltd"
        )
        org_formatted = org_name if org_name.startswith("Organization:") \
                        else f"Organization: {org_name}"

        spdx_doc = {
            "spdxVersion": metadata.get("spdxVersion", "SPDX-2.3"),
            "dataLicense": metadata.get("dataLicense", "CC0-1.0"),
            "SPDXID": metadata.get("SPDXID", "SPDXRef-DOCUMENT"),
            "name": name,
            "documentNamespace": f"urn:uuid:{uuid.uuid4()}",
            "creationInfo": {
                "creators": [org_formatted],
                "created": spdx_timestamp(),
            },
            "documentDescribes": [ROOT_PKG] if ROOT_PKG in all_package_ids else [],
            "packages": packages,
            "relationships": relationships,
        }

        return spdx_doc, vbf_part_number

    except Exception as e:
        import traceback
        print(f"  ERROR processing {file_path}: {e}")
        traceback.print_exc()
        return None, None


def save_sbom_to_json(spdx_doc: dict, output_path: str) -> bool:
    """Serialize SPDX document to a JSON file."""
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(spdx_doc, f, indent=2, ensure_ascii=False)
        size_kb = os.path.getsize(output_path) / 1024
        print(f"  Saved: {output_path} ({size_kb:.1f} KB)")
        return True
    except Exception as e:
        print(f"  ERROR saving {output_path}: {e}")
        return False


def _output_filename(spdx_doc: dict, vbf_part_number: str, excel_path: str) -> str:
    """Determine the output JSON filename."""
    if vbf_part_number:
        base = vbf_part_number
    else:
        doc_name = spdx_doc.get("name", "")
        if "VBF-" in doc_name:
            base = clean_vbf_part_number(doc_name.split("VBF-")[1].split(" ")[0])
        else:
            base = clean_vbf_part_number(os.path.splitext(os.path.basename(excel_path))[0])

    filename = base if base.endswith(".sbom.json") else f"{base}.sbom.json"
    return filename


def process_single_file(excel_file: str, output_dir: str = "."):
    """Process one Excel file → JSON."""
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        return False, None

    spdx_doc, vbf_part_number = extract_sbom_from_excel(excel_file)
    if not spdx_doc:
        return False, None

    output_file = _output_filename(spdx_doc, vbf_part_number, excel_file)
    os.makedirs(output_dir, exist_ok=True)
    success = save_sbom_to_json(spdx_doc, os.path.join(output_dir, output_file))
    return success, output_file


def process_folder(folder_path: str, file_pattern: str = "*.xlsx", output_dir: str = "."):
    """Process all matching Excel files in a folder."""
    if not os.path.exists(folder_path):
        print(f"ERROR: Folder not found: {folder_path}")
        return 0, [], []

    excel_files = glob.glob(os.path.join(folder_path, file_pattern))
    if not excel_files:
        for ext in ("*.xls", "*.xlsm", "*.xlsx"):
            excel_files.extend(glob.glob(os.path.join(folder_path, ext)))
        excel_files = list(set(excel_files))

    if not excel_files:
        print(f"ERROR: No Excel files found in {folder_path}")
        return 0, [], []

    print(f"Found {len(excel_files)} Excel file(s):")
    for i, f in enumerate(sorted(excel_files), 1):
        print(f"  {i}. {os.path.basename(f)}")

    success_count, failed_files, generated_files = 0, [], []

    print("\nProcessing files...")
    for excel_file in sorted(excel_files):
        print(f"\n--- {os.path.basename(excel_file)} ---")
        try:
            ok, out_file = process_single_file(excel_file, output_dir)
            if ok:
                success_count += 1
                generated_files.append(out_file)
            else:
                failed_files.append(os.path.basename(excel_file))
        except Exception as e:
            print(f"  FAILED: {e}")
            failed_files.append(os.path.basename(excel_file))

    return success_count, failed_files, generated_files


def _print_summary(output_dir: str, generated_files: list, total: int,
                   success: int, failed: list):
    sep = "=" * 60
    print(f"\n{sep}")
    print("Done!")
    print(sep)
    print(f"Total : {total}")
    print(f"OK    : {success}")
    if failed:
        print(f"Failed: {len(failed)}")
        for f in failed[:10]:
            print(f"  - {f}")

    for json_file in sorted(generated_files):
        path = os.path.join(output_dir, json_file)
        if not os.path.exists(path):
            continue
        size_kb = os.path.getsize(path) / 1024
        print(f"\n  {json_file} ({size_kb:.1f} KB)")
        try:
            with open(path, encoding='utf-8') as f:
                data = json.load(f)
            pkgs = data.get('packages', [])
            rels = data.get('relationships', [])
            contains = [r for r in rels if r['relationshipType'] == 'CONTAINS']
            depends  = [r for r in rels if r['relationshipType'] == 'DEPENDS_ON']
            has_vbf  = any(p['SPDXID'] == 'SPDXRef-Package-VBF' for p in pkgs)
            print(f"    Packages      : {len(pkgs)}")
            print(f"    CONTAINS      : {len(contains)}")
            print(f"    DEPENDS_ON    : {len(depends)}")
            if has_vbf:
                expected = len(pkgs) - 1
                status = "✅" if len(contains) == expected else "⚠️"
                print(f"    VBF check     : {status} ({len(contains)}/{expected})")
        except Exception:
            pass
    print(sep)


def main():
    parser = argparse.ArgumentParser(
        description="Convert automotive SBOM Excel files to SPDX 2.3 JSON.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sbom_converter.py .                   # convert all .xlsx in current folder
  python sbom_converter.py SBOM-APP.xlsx       # convert a single file
  python sbom_converter.py ./input -o ./output # specify output directory
        """
    )
    parser.add_argument("input", nargs="?", default=".",
                        help="Excel file or folder path (default: current directory)")
    parser.add_argument("-p", "--pattern", default="*.xlsx",
                        help="File glob pattern (default: *.xlsx)")
    parser.add_argument("-o", "--output", default=".",
                        help="Output directory (default: current directory)")
    args = parser.parse_args()

    print("=" * 60)
    print("SBOM Excel → SPDX 2.3 JSON Converter")
    print("=" * 60)

    input_path = args.input
    output_dir = args.output

    if os.path.isdir(input_path):
        success, failed, generated = process_folder(input_path, args.pattern, output_dir)
        total = len(glob.glob(os.path.join(input_path, args.pattern)))
    elif os.path.isfile(input_path):
        if not input_path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
            print(f"ERROR: Not an Excel file: {input_path}")
            sys.exit(1)
        ok, out_file = process_single_file(input_path, output_dir)
        success, failed, generated = (1, [], [out_file]) if ok else (0, [input_path], [])
        total = 1
    else:
        print(f"ERROR: Path does not exist: {input_path}")
        sys.exit(1)

    _print_summary(output_dir, generated, total, success, failed)


if __name__ == "__main__":
    main()
